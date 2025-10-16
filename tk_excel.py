# -*- coding:utf-8 -*-
import datetime
import tkinter as tk
from tkinter.ttk import *
from tkinter import *
from tkinter import filedialog
import threading
import unicodedata
import re
from urllib.parse import urlparse
from tkinter import scrolledtext
import pandas  as pd
from openpyxl.styles import Alignment, Font, Color, PatternFill
from PIL import Image, ImageTk, ImageDraw, ImageFont, ImageEnhance, ImageFilter
from io import BytesIO
from tkcalendar import DateEntry
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image  as OpenpyxlImage
import base64
import json, requests, os, io
from openpyxl import Workbook
from collections import Counter
import configparser
import difflib
import qrcode
from pyzbar.pyzbar import decode
import random
import sys
import os
from pathlib import Path
import keyboard
import win32clipboard
import win32con
import shutil
import struct
import pyperclip
import gc


class Config(object):
    def __init__(self, config_filename="config\config.ini"):
        file_path = os.path.join(os.getcwd(), config_filename)
        self.cf = configparser.ConfigParser()
        self.cf.read(file_path)
        self.config_file = file_path

    def get_sections(self):
        return self.cf.sections()

    def get_options(self, section):
        return self.cf.options(section)

    def get_content(self, section):
        result = {}
        for option in self.get_options(section):
            value = self.cf.get(section, option)
            result[option] = int(value) if value.isdigit() else value
        return result

    def set_option(self, section, option, value):
        """设置配置项"""
        if not self.cf.has_section(section):
            self.cf.add_section(section)
        self.cf.set(section, option, str(value))

    def remove_option(self, section, option):
        """删除配置项"""
        if self.cf.has_section(section):
            self.cf.remove_option(section, option)
            if not self.cf.options(section):
                self.cf.remove_section(section)

    def save_config(self):
        """保存配置到文件"""
        with open(self.config_file, 'w') as configfile:
            self.cf.write(configfile)


def strGetlen(strn, strx, strend):
    """
    截取字符串
    :param strn:
    :param strx:
    :param strend:
    :return:
    """
    sint = strn.find(strx)
    strn = strn[sint + len(strx):]
    if strend == "":
        return strn
    eint = strn.find(strend)
    return strn[:eint]


# 将函数打包进线程
def thread_is(func):
    '''将函数打包进线程'''
    # 创建
    t = threading.Thread(target=func)
    # 守护 !!!
    # t.setDaemon(True)
    # 启动
    t.start()
    # 阻塞--卡死界面！
    # t.join()


def get_date():
    """
    获取当前时间
    :return:
    """

    return datetime.datetime.now()


def is_image(filename):
    try:
        Image.open(filename)
        return True
    except IOError:
        return False


def is_empty(var):
    """通用空值检测（支持字符串/列表/字典等）"""
    if var is None:
        return True
    if isinstance(var, (str, list, dict, set)):
        return not bool(var)
    return False


class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.conf = Config()
        self.content_text_label = []
        self.verify = False
        self.data_init()
        self.line_num = 0
        self.callbck_see = True
        self.search_see = True
        self.file_text = None
        self.folder_config = None
        self.folder_img = None
        self.master = master
        self.pack()
        self.window_init()
        self.create_notebook()

    def window_init(self):
        width, height = self.master.maxsize()  # 获取窗体大小
        self.master.title("excel工具")  # 标题
        if width < 1152:
            width, height = 1080, 600
        self.width = 500
        self.height = 400
        widths = int(width / 2.5)  # 屏幕显示位置
        heights = int(height / 2.5)  # 屏幕显示位置
        self.master.geometry("{}x{}+{}+{}".format(self.width, self.height, widths, heights))  # 设置大小
        ico = open('config\login.ico', 'rb').read()
        image_data = BytesIO(ico)
        img = Image.open(image_data)
        photo = ImageTk.PhotoImage(img)
        self.master.iconphoto(True, photo)
        # 绑定窗口变动事件
        self.master.bind('<Configure>', self.WindowResize)
        self.master.resizable(0, 0)
        # -*- encoding=utf-8 -*-
        self.frame = Frame(
            master=self.master,  # 父容器
            # bg='yellow',  # 背景颜色
            relief='groove',  # 边框的3D样式 flat、sunken、raised、groove、ridge、solid。
            bd=0,  # 边框的大小
            height=100,  # 高度
            width=100,  # 宽度
            padx=0,  # 内间距，字体与边框的X距离
            pady=0,  # 内间距，字体与边框的Y距离
            cursor='arrow',  # 鼠标移动时样式 arrow, circle, cross, plus...
        )

    def ocr_data(self, base64_str, format='text'):
        url = self.text_entry.get()
        data = {
            "base64": base64_str,
            # 可选参数示例
            "options": {
                "data.format": format,
            }
        }
        headers = {"Content-Type": "application/json"}
        data_str = json.dumps(data)
        response = requests.post(url, data=data_str, headers=headers, verify=False)
        res_dict = json.loads(response.text)
        if res_dict['code'] == 100:
            self.conf.set_option("Setting", "url", url)
            self.conf.save_config()
            return res_dict
        else:
            return {'data': ""}

    def create_rounded_icon(self, icon_name, output_size=100, radius=8):
        # 1. 加载并预处理图标（2倍超采样）
        icon = Image.open(icon_name).convert("RGBA")
        icon = icon.resize((output_size * 2, output_size * 2), Image.ANTIALIAS)

        # 2. 创建超采样遮罩
        mask = Image.new('L', (output_size * 2, output_size * 2), 0)
        draw = ImageDraw.Draw(mask)

        # 3. 绘制圆角路径（使用新版rounded_rectangle方法）
        if hasattr(ImageDraw, 'rounded_rectangle'):  # Pillow 9.1+
            draw.rounded_rectangle([0, 0, output_size * 2, output_size * 2],
                                   radius=radius * 2, fill=255)
        else:  # 兼容旧版
            draw.pieslice((0, 0, radius * 4, radius * 4), 180, 270, fill=255)
            draw.pieslice((output_size * 2 - radius * 4, 0, output_size * 2, radius * 4), 270, 360, fill=255)
            draw.pieslice((0, output_size * 2 - radius * 4, radius * 4, output_size * 2), 90, 180, fill=255)
            draw.pieslice((output_size * 2 - radius * 4, output_size * 2 - radius * 4,
                           output_size * 2, output_size * 2), 0, 90, fill=255)
            draw.rectangle([radius * 2, 0, output_size * 2 - radius * 2, output_size * 2], fill=255)
            draw.rectangle([0, radius * 2, output_size * 2, output_size * 2 - radius * 2], fill=255)

        # 4. 下采样并应用抗锯齿
        mask = mask.resize((output_size, output_size), Image.ANTIALIAS)

        # 5. 边缘羽化（可选）
        mask = mask.filter(ImageFilter.GaussianBlur(radius=0.7))

        # 6. 应用遮罩并下采样图标
        icon = icon.resize((output_size, output_size), Image.ANTIALIAS)
        icon.putalpha(mask)

        return icon

    def qrcode_make(self, data):
        # 生成二维码
        # data = "https://work.weixin.qq.com/ca/cawcdeb8dbcb42d76a"
        qr = qrcode.QRCode(
            # 尝试提高版本号，提供更多数据存储和纠错空间
            version=1,
            # 使用更高的纠错级别
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=0,
        )
        qr.add_data(data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        # 获取二维码的尺寸
        width, height = qr_img.size
        box_size = qr.box_size

        # 打开 Logo 图片
        try:
            logo_path = os.path.join(self.folder_config, "wechat_logo.png")
            logo = Image.open(logo_path)
            logo = logo.convert("RGBA")  # 确保 Logo 是 RGBA 格式
        except FileNotFoundError:
            print("未找到指定的 Logo 图片，请检查文件路径和文件名。")
            exit(1)

        # 创建一个新的空白图像
        new_img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(new_img)

        # 进一步减小间隙大小
        gap = 2

        # 定位图案的大小（固定为 7 个模块）
        pattern_size = 7 * box_size

        # 进一步缩小 logo 区域的大小（这里设置为二维码宽度的 1/6）
        logo_size = width // 4
        # 计算 logo 区域的左上角坐标
        logo_x = (width - logo_size) // 2
        logo_y = (height - logo_size) // 2

        top_bottom = None

        # 遍历二维码的每个像素
        for y in range(0, height, box_size):
            for x in range(0, width, box_size):
                # 检查是否在左上角、右上角和左下角的定位图案及对应行内
                in_top_left_row = y < pattern_size
                in_top_right_row = y < pattern_size
                in_bottom_left_row = y > height - pattern_size
                in_top_left = x < pattern_size and in_top_left_row
                in_top_right = x > width - pattern_size and in_top_right_row
                in_bottom_left = x < pattern_size and in_bottom_left_row
                if qr_img.getpixel((x, y)) == 0:  # 如果是黑色像素
                    if in_bottom_left:
                        top_bottom = x, y
        top_left = top_bottom[0] + 10
        top_right = top_bottom[1] - 70
        bottom_left = top_bottom[0] + 50
        # 遍历二维码的每个像素
        for y in range(0, height, box_size):
            for x in range(0, width, box_size):
                # 检查是否在左上角、右上角和左下角的定位图案及对应行内
                in_top_left_row = y < pattern_size
                in_top_right_row = y < pattern_size
                in_bottom_left_row = y > height - pattern_size
                in_top_left = x < pattern_size and in_top_left_row
                in_top_right = x > width - pattern_size and in_top_right_row
                in_bottom_left = x < pattern_size and in_bottom_left_row

                # 检查是否在 logo 区域内
                in_logo_area = logo_x <= x < logo_x + logo_size and logo_y <= y < logo_y + logo_size

                if qr_img.getpixel((x, y)) == 0:  # 如果是黑色像素

                    if x < top_left and y < top_left:
                        draw.rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    elif x > top_right and y < top_left:
                        draw.rounded_rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    elif x < top_left and y > top_right:
                        draw.rounded_rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    # if x < 113 and y < 113:
                    #     draw.rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    # elif x > 370 and y < 113:
                    #     draw.rounded_rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    # elif x < 113 and y > 370:
                    #     draw.rounded_rectangle([(x, y), (x + box_size, y + box_size)], fill='black')
                    elif not in_logo_area:
                        # 如果不在 logo 区域内，绘制带有间隙的小正方形
                        draw.rounded_rectangle([(x + gap, y + gap), (x + box_size - gap, y + box_size - gap)],
                                               fill='black',
                                               radius=gap)

        # 进一步缩小 Logo 尺寸
        factor = 5
        size_w = int(new_img.size[0] / factor)
        size_h = int(new_img.size[1] / factor)
        logo = logo.resize((size_w, size_h), Image.ANTIALIAS)

        # 计算 Logo 在二维码中的位置，使其居中
        w = int((new_img.size[0] - logo.size[0]) / 2) + 5
        h = int((new_img.size[1] - logo.size[1]) / 2) + 5

        # 将 Logo 嵌入二维码
        new_img.paste(logo, (w, h), logo)
        return new_img

    def data_init(self):
        self.channel_list = []
        self.price_count = 0
        self.complete_number = 0
        self.repair_number = 0
        self.otehr_number = 0
        self.boot_dict = {}
        self.status_dict = {}
        self.folder_selected = None
        self.qr_folder = None

    def work_book(self, file_path, table_name):
        try:
            # 加载工作簿和工作表
            wb = load_workbook(file_path, read_only=True)
            sheet = wb[table_name]
            all_dict = {}
            all_data = []
            dict_index = 0
            for row in sheet.iter_rows():
                row_data = []
                index = 0
                for cell in row:
                    if dict_index == 0:
                        all_dict.update({cell.value: []})
                    else:
                        if index >= len(all_dict):
                            continue
                        # 固定列 转小写
                        if type(cell.value) == str and ("C" in cell.coordinate or "F" in cell.coordinate):
                            value_res = str(cell.value).lower()
                            value_res = value_res.replace(" ", '')
                            if '吉祥' in value_res:
                                value_res = '吉祥'
                            if 'kk' in value_res:
                                value_res = 'kk'
                        else:
                            value_res = cell.value
                        all_dict[list(all_dict)[index]].append(value_res)
                    row_data.append(cell.value)
                    index += 1
                all_data.append(row_data)
                dict_index += 1
            for itme in all_dict:
                data_list = pd.Series(all_dict[itme])
                all_dict[itme] = data_list
            df = pd.DataFrame(all_dict)
            return df
        except openpyxl.utils.exceptions.InvalidFileException:
            self.insert_text("文件无效或格式不支持", 'red')
        except Exception as e:
            self.insert_text(f"{e}", 'red')

    def _remove_invalid_unicode(self, input_str):
        # 定义 Unicode 字符串
        result = ""
        for char in str(input_str):
            try:
                unicodedata.name(char)
                pattern = re.compile("[^\u0000-\uFFFF]")
                replacement = "replacement"
                char = pattern.sub(replacement, char)
                result += char
            except ValueError:
                pass
        return result

    def verify_see(self, event, verify_see, verify):
        if verify:
            if verify_see:
                self.callbck_see = False
            else:
                self.callbck_see = True
        else:
            if verify_see:
                self.search_see = False
            else:
                self.search_see = True

    def insert_text(self, text, tag=None, label=None):
        self.line_num += 1
        if self.line_num >= 300:
            # self.text_callbck_msg.delete('1.0', tk.END)
            self.text_callbck_msg.delete('1.0', '100.0')
            self.line_num -= 100
            self.text_callbck_msg.update()
        # text = self._remove_invalid_unicode(text)
        self.text_callbck_msg.insert(tk.END, '{}  \n'.format(text))
        if tag != None:
            if label != None:
                label = "{}".format(label)
                index = text.find(label)
                start_index = "{}.{}".format(self.line_num, index)
                end_index = "{}.{}".format(self.line_num, index + len(label))
                self.text_callbck_msg.tag_add(tag, start_index, end_index)
                self.text_callbck_msg.tag_config(tag, foreground=tag)
            else:
                start_index = f"{self.line_num}.0"  # 获取当前行的第一个字符的索引
                end_index = f"{self.line_num}.end"  # 获取当前行的最后一个字符的索引
                self.text_callbck_msg.tag_add(tag, start_index, end_index)  # 添加标签“my_tag” (索引从1开始)
                self.text_callbck_msg.tag_config(tag, foreground=tag)  # 将标签“my_tag”设置为红色  foreground=""文字颜色
        if self.callbck_see:
            self.text_callbck_msg.yview_moveto(1.0)

    def notebook_size(self):
        self.frame.place(width=self.width, height=self.height)
        self.TabStrip.place(x=0, y=0, width=self.width, height=self.height / 1.6)
        self.TabStrip1.place(x=0, y=self.height / 1.6, width=self.width, height=self.height / 1.3)
        self.text_callbck_msg.place(x=10, y=0, width=(self.width - 10), height=int(self.height / 1.6))

        self.tab_control.pack(fill=BOTH, expand=True)
        self.text_date.place(x=120, y=10)
        self.cal.place(x=170, y=10)
        self.start_but.place(x=120, y=40)
        self.loading_but.place(x=120 + 115, y=40)
        self.label_port.place(x=60, y=10)
        self.text_entry.place(x=60 + 115, y=10)
        self.start_but2.place(x=20, y=40)
        self.start_but3.place(x=20 + + (115 * 1), y=40)
        self.loading_but2.place(x=20 + (115 * 3), y=40)
        self.start_link_but1.place(x=20 + (115 * 2), y=40)
        self.text_date1.place(x=60, y=90)

        self.start_img_but.place(x=60, y=30)
        self.start_link_but.place(x=60 + (115 * 1), y=30)
        self.text_entry1.place(x=60 + (115 * 1), y=90)

        self.start_img_but1.place(x=60, y=60)

        self.start_img_but2.place(x=60 + (115 * 1), y=60)

        self.generate_but.place(x=60 + (115 * 2), y=60)

        self.start_config_but.place(x=60 + (115 * 2), y=30)

        self.generate1_but.place(x=60 + (115 * 2), y=90)
        self.get_img_but.place(x=60 + (115 * 1), y=60)
        config_dict = self.conf.get_content("Setting")
        self.text_entry.insert(0, config_dict["url"])
        self.text_entry1.insert(0, config_dict["fission"])

    # 窗口尺寸调整处理函数
    def WindowResize(self, event):
        new_width = self.master.winfo_width()
        new_height = self.master.winfo_height()
        if new_width == 1 and new_height == 1:
            return
        if self.width != new_width or self.height != new_height:
            self.width = new_width
            self.height = new_height
            self.notebook_size()

    def move_files_to_parent_folder(self, parent_folder):
        current_dir_list = []
        # 遍历父文件夹下的所有文件和文件夹
        for root, dirs, files in os.walk(parent_folder):
            for name in files:
                try:
                    if root == parent_folder:
                        continue
                    # 构建文件的完整路径
                    file_path = os.path.join(root, name)
                    # 检查目标路径下是否存在同名文件
                    target_path = os.path.join(parent_folder, os.path.basename(file_path))
                    if os.path.exists(target_path):
                        os.remove(target_path)  # 先删除目标文件
                    # 将文件移动到父文件夹
                    shutil.move(file_path, parent_folder)
                except Exception as e:
                    self.insert_text(f'移动文件夹错误:{e}', 'red')
                # 移动完所有删除文件夹
                if not os.listdir(root):
                    os.rmdir(root)
            if root not in current_dir_list:
                current_dir_list.append(root)
        current_dir_list.reverse()
        for current_dir in current_dir_list:
            try:
                if parent_folder == current_dir:
                    continue
                # 移动完所有删除文件夹
                if not os.listdir(current_dir):
                    os.rmdir(current_dir)
            except Exception as e:
                self.insert_text(f'删除文件夹错误:{e}', 'red')

    def start_select(self, tab):
        if tab == 1:
            self.file_path = filedialog.askopenfilename()
            if self.file_path.lower().endswith('.xlsx'):
                self.insert_text(self.file_path, 'green')
                self.verify = True
                df, price_df = self.workbook_init(self.file_path)
                if is_empty(df) == False:
                    # 确保'Date'列是日期格式
                    df['日期'] = pd.to_datetime(df['日期'])
                    crruent_time = max(list(df['日期'])) if list(df['日期']) else None
                    # 选择文件设置默认时间
                    self.cal.set_date(crruent_time)
            else:
                self.insert_text("请选择xlsx文件", 'red')
        elif tab == 2:
            self.folder_selected = filedialog.askdirectory()
            # 检查用户是否选择了文件夹
            if self.folder_selected:
                self.insert_text(f"选中的文件夹:{self.folder_selected}", 'green')
            else:
                self.insert_text("没有选择文件夹", 'red')
            # self.data_path = filedialog.askopenfilename()
            # if self.data_path.lower().endswith('.xls'):
            #     base_dirname = os.path.dirname(self.data_path)
            #     base_name = os.path.basename(self.data_path)
            #     copy_path = os.path.join(base_dirname, f"{base_name.split('.')[0]}复制.xlsx")
            #     with open(self.data_path, 'rb') as src_file, open(copy_path, 'wb') as dest_file:
            #         shutil.copyfileobj(src_file, dest_file)
            #     self.data_path = copy_path
        elif tab == 3:
            self.qr_folder = filedialog.askdirectory()
            # 检查用户是否选择了文件夹
            if self.qr_folder:
                self.insert_text(f"选中的文件夹:{self.qr_folder}", 'green')
            else:
                self.insert_text("没有选择文件夹", 'red')
        elif tab == 4:
            self.file_text = filedialog.askopenfilename()
            if self.file_text.lower().endswith('.txt'):
                self.insert_text(self.file_text, 'green')
            else:
                self.insert_text("请选择txt文件", 'red')
        elif tab == 5:
            pass
            # self.folder_config = filedialog.askdirectory()
            # # 检查用户是否选择了文件夹
            # if self.folder_config:
            #     self.insert_text(f"选中的文件夹:{self.folder_config}", 'green')
            # else:
            #     self.insert_text("没有选择文件夹", 'red')
        elif tab == 6:
            self.folder_img = filedialog.askdirectory()
            # 检查用户是否选择了文件夹
            if self.folder_img:
                self.insert_text(f"选中的文件夹:{self.folder_img}", 'green')
                self.file_text = None
            else:
                self.insert_text("没有选择文件夹", 'red')

    def file_operate(self, tab, img_file, new_name=None):
        if is_image(img_file):
            if tab == 1:
                try:
                    os.rename(img_file, new_name)
                    # shutil.move(img_file, new_name)
                except Exception as e:
                    self.insert_text(f'{e}', 'red')
            elif tab == 2:
                image = Image.open(img_file)
                decoded_objects = decode(image)
                if decoded_objects:
                    url = decoded_objects[0].data.decode('utf-8')
                    result = self.qrcode_img(url, img_file)
                    if result:
                        return True

    def start_img(self, tab):
        try:
            if self.folder_img:
                folder_list = os.listdir(self.folder_img)
                img_count = 0
                verify_flie = False
                for item in folder_list:
                    img_path = os.path.join(self.folder_img, item)
                    if is_image(img_path) or verify_flie:
                        verify_flie = True
                        if tab == 1:
                            img_basename = os.path.basename(self.folder_img)
                            if '-' in item:
                                match = re.search(img_basename + "\d+", item)
                                if img_basename in item.split("-")[0]:
                                    if match:
                                        self.insert_text(f'{img_basename}已修改', 'green')
                                        break
                            img_count += 1
                            new_name = os.path.join(self.folder_img, f"{img_basename}{img_count}-{item}")
                            self.file_operate(tab, img_path, new_name)
                        elif tab == 2:
                            img_count += 1
                            result = self.file_operate(tab, img_path)
                            if result:
                                return
                        continue
                    if os.path.isdir(img_path):
                        self.move_files_to_parent_folder(img_path)
                        img_list = os.listdir(img_path)
                        count = 0
                        for img in img_list:
                            img_file = os.path.join(img_path, img)
                            if os.path.isdir(img_file):
                                continue
                            if is_image(img_file):
                                if tab == 1:
                                    if '-' in img:
                                        match = re.search(item + "\d+", img)
                                        if item in img.split("-")[0]:
                                            if match:
                                                self.insert_text(f'{item}已修改', 'green')
                                                break
                                            elif re.search(item + "-\d+", img):
                                                self.insert_text(f'{item}已修改', 'green')
                                                break
                                    count += 1
                                    img_count += 1
                                    new_name = os.path.join(img_path, f"{item}-{count}-{img}")
                                    self.file_operate(tab, img_file, new_name)
                                elif tab == 2:
                                    result = self.file_operate(tab, img_file)
                                    if result:
                                        return
                                    img_count += 1
                                    count += 1
                        self.insert_text(f'{item}:{count}', 'green')
                self.insert_text(f'共计:{img_count}', 'green')
                os.startfile(self.folder_img)
            else:
                self.insert_text("没有图片选择文件夹", 'red')
                self.insert_text("请选择父文件夹", 'red')
        except Exception as e:
            self.insert_text(f'{e}', 'red')

    def team_dict(self, boot_dict):
        team_list = []
        for team_itme in boot_dict:
            for item in boot_dict[team_itme].keys():
                item = item.replace(" ", '')
                if item in team_list:
                    continue
                team_list.append(item)
        team_list_cale = []
        for team_itme in team_list:
            item = team_itme.replace(" ", '')
            if re.match(r'\d+[\(\d+\)|（\d+）]', item):
                team = re.sub(r'\(\d+\)', '', item)
                team = re.sub(r'（\d+）', '', team)
            else:
                team = re.sub(r'\d+', '', item)
            team_list_cale.append(team)
        # 创建一个空字典来存储计数
        count_dict = {}
        for char in team_list_cale:
            if char in count_dict:
                count_dict[char] += 1
            else:
                count_dict[char] = 1
        return count_dict

    def start_loadning(self):
        try:
            if self.verify:
                self.data_init()
                self.work_result(self.file_path)
                team_dict = self.team_dict(self.boot_dict)
                group_num = 0
                for team_item in team_dict:
                    self.insert_text(f'{team_item}：{team_dict[team_item]}组')
                    group_num += team_dict[team_item]
                esc_dict = self.team_dict(self.status_dict)
                for team_item in esc_dict:
                    self.insert_text(f'{team_item}：退:{esc_dict[team_item]}组')
                text_red = f'用:{self.complete_number}/补:{self.repair_number}/其他:{self.otehr_number}'
                for itme in self.channel_list:
                    self.insert_text(itme[0], 'green', itme[1])
                self.insert_text(
                    f'共计：{self.complete_number + self.repair_number + self.otehr_number}个 {group_num}组 {text_red}',
                    'red', text_red)
                self.insert_text(f'====================共花费:{self.price_count}====================', 'green')
                # 提取数字字符的机器
                numeric_machines = [x for x in self.boot_dict.keys() if x.isdigit() or x.split('号')[0].isdigit()]
                # 提取非数字字符的机器
                non_numeric_machines = [x for x in self.boot_dict.keys() if
                                        not (x.isdigit() or x.split('号')[0].isdigit())]
                # 将数字 字符进行排序
                boot_list = sorted(numeric_machines, key=lambda x: int(''.join(filter(str.isdigit, x))))
                # 把非数字字符添加到列表
                boot_list += non_numeric_machines
                for boot in boot_list:
                    one = 0
                    twelve = 0
                    for item in self.boot_dict[boot]:
                        if self.boot_dict[boot][item] >= 118:
                            one += 1
                        else:
                            twelve += 1
                    self.insert_text(f'{boot}:{len(self.boot_dict[boot].keys())}组\t\t\t12点:{twelve}组 1点:{one}组')
                self.insert_text(f'共：{group_num}组')
            else:
                self.insert_text("请选择xlsx文件", 'red')
        except Exception as e:
            self.insert_text(f"{e}", 'red')

    def start_work(self):
        try:
            if self.folder_selected == None:
                self.insert_text("未选择出粉文件夹", 'red')
                return
            # 创建一个Workbook对象，这将会创建一个新的Excel文件
            wb = Workbook()
            # 获取当前活跃的worksheet
            ws = wb.active
            # 更改worksheet的名称
            ws.title = "出粉数据"
            # 添加表头
            ws.append(["昵称", "广告码ID", "出粉ID", "粉数", '访问量', '准确率', "是否存在广告码"])
            for column in "ABCDEFG":  # 根据需要调整列宽和行高，例如A、B、C列
                ws.column_dimensions[column].width = 30  # 设置列宽为20个字符宽
                # 设置第一列的背景颜色为淡黄色（FFEB9C）
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                ws.cell(row=1, column=1).fill = fill
                font = Font(color=Color(rgb='000000'))  # 字体颜色
                cell = ws[column][0]
                cell.font = font
                cell.fill = fill
            self.cawcde_num = 0
            if self.qr_folder:
                qr_folder = self.qr_folder
                qr_list = os.listdir(qr_folder)
                qr_file = [itme.split('-')[1].split('.')[0] for itme in qr_list]
            elif self.file_text:
                cawcdeb_list = open(self.file_text, 'r').readlines()
                qr_list = [item.replace("\n", "") for item in cawcdeb_list]
                qr_file = [itme.split('/')[-1] for itme in qr_list]
            else:
                self.insert_text("未选择广告码文件夹", 'red')
                qr_folder = os.path.join(self.folder_selected, "广告码")
                self.insert_text(f"默认选择:{qr_folder}", 'green')
                try:
                    qr_list = os.listdir(qr_folder)
                    qr_file = [itme.split('-')[1].split('.')[0] for itme in qr_list]
                except FileNotFoundError as e:
                    self.insert_text(f"{e}", 'red')
                    qr_list = []
                    qr_file = []

            # 解析图片 提取粉数 os.path.dirname()
            result = self.img_data(self.folder_selected)
            # self.cawcdeb_list = result.keys()
            num_count = 0
            exist_id_num = 0
            no_id_num = 0
            id_num = 0
            no_img_num = 0
            adver_list = []
            cawcdeb_in_list = []
            # 使用列表推导式找到两个列表中的相同元素
            column_values = [element for element in qr_file if element in result.keys()]
            for exist_itme in qr_file:
                similar_words = difflib.get_close_matches(exist_itme, result.keys())
                ratio_dict = {}
                if similar_words:
                    for code_id in similar_words:
                        d = difflib.SequenceMatcher(None, exist_itme, code_id)
                        ratio_dict.update({code_id: d.ratio() * 100})
                    max_key = max(ratio_dict, key=ratio_dict.get)
                    ratio = ratio_dict[max_key]
                    if ratio >= 100:
                        pass
                    elif ratio >= 70:
                        if max_key in adver_list:
                            continue
                        elif max_key in column_values:
                            continue
                    else:
                        continue
                    ws.append(
                        [result[max_key]["name"], exist_itme, max_key, result[max_key]["num"],
                         result[max_key]["link_num"],
                         f'{int(ratio)}%', '是'])
                    exist_id_num += result[max_key]["num"]
                    id_num += 1
                    adver_list.append(max_key)
                    cawcdeb_in_list.append(exist_itme)
            for itme in result:
                name = result[itme]["name"]
                num = result[itme]["num"]
                link_num = result[itme]["link_num"]
                num_count += num
                if itme in adver_list:
                    continue
                ws.append([name, '', itme, num, link_num, '0%', "否"])
                no_id_num += 1
            for elem in qr_file:
                if elem in cawcdeb_in_list:
                    continue
                no_img_num += 1
                ws.append(["", elem, "", '', '', '0%', "是"])

            for row in ws.iter_rows():
                row_value = row[5].value
                if row_value == '':
                    continue
                if row[5].value[:-1].isdigit() == False:
                    continue
                ratio = int(row[5].value[:-1])
                if ratio == 0:
                    continue
                if ratio >= 100:
                    color = '90EE90'
                elif ratio >= 80:
                    color = 'FFFFE0'
                else:
                    color = 'FF0000'
                fill = PatternFill(start_color=color, end_color="FFCCCC", fill_type="solid")
                row[1].fill = fill
                row[2].fill = fill
                row[5].fill = fill
            ws.append(["", ""])
            ws.append(["", "出粉ID总数:", self.cawcde_num])
            ws.append(["", "广告码总数:", len(qr_list)])
            ws.append(["", "广告码对应出粉总数:", id_num])
            ws.append(["", "未找到广告码总数:", no_id_num])
            ws.append(["", "未找到出粉图总数:", no_img_num])
            ws.append(["", ""])
            ws.append(["", "广告码合计:", '', exist_id_num])
            ws.append(["", "总粉合计:", '', num_count])
            ws.append(["", ""])
            # 使用Counter统计
            cawcdeb_count = dict(Counter(self.cawcdeb_list))
            ws.append(["", "", f"重复粉ID:"])
            repeat_num = 0
            for item in cawcdeb_count:
                coun = cawcdeb_count[item]
                if coun > 1:
                    repeat_num += 1
                    ws.append(["", "", item])
            ws.append(["", "共计:", repeat_num])
            # 设置所有单元格文字居中
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            adver_path = os.path.join(os.path.dirname(self.folder_selected),
                                      f'{os.path.basename(self.folder_selected)}出粉记录.xlsx')
            self.insert_text(adver_path, 'green')
            # 保存文件
            wb.save(adver_path)
            os.startfile(os.path.dirname(self.folder_selected))

        except Exception as e:
            self.insert_text(f"{e}", 'red')

    def start_get_img(self):
        try:
            headers = {
                "Content-Type": "text/html; charset=utf-8",
                "User-Agent": 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1',

            }
            url = 'https://x.wobew.com/app/index.php?i=2&c=entry&search=1&do=dataJson&m=yy_shequn2'
            result = requests.get(url, headers=headers, verify=False).json()
            # 获取当前执行的 exe 文件的路径
            exe_path = sys.executable
            # 获取 exe 文件所在的目录
            exe_dir = os.path.dirname(exe_path)
            downloads = os.path.join(exe_dir, 'Downloads')
            os.makedirs(downloads, exist_ok=True)
            data_time = get_date().strftime('%Y-%m-%d')
            data_path = os.path.join(downloads, data_time)
            os.makedirs(data_path, exist_ok=True)
            self.insert_text(f'图片下载中...', 'green')
            os.startfile(data_path)
            count = 0
            for itme in result:
                count += 1
                headers["Content-Type"] = 'image/jpeg'
                headers["Referer"] = 'https://x.wobew.com/'
                img_url = f"https://shequn-1255833192.cos.ap-nanjing.myqcloud.com/{itme['qrQunImg']}"
                img_result = requests.get(img_url, headers=headers, verify=False)
                img_file = os.path.join(data_path, f'{itme["id"]}.png')
                with open(img_file, 'wb') as f:
                    f.write(img_result.content)
            self.insert_text(f'保存地址:{data_path}', 'green')
            self.insert_text(f'共计:{count}张', 'green')
        except Exception as e:
            self.insert_text(f'{e}', 'red')

    def create_notebook(self):
        self.TabStrip = Notebook(self.frame)
        self.text_callbck_msg = scrolledtext.ScrolledText(self.TabStrip, font=("Times New Roman", 10))
        self.scroll = tk.Scrollbar(self.TabStrip, command=self.text_callbck_msg.yview)
        self.scroll.pack(side=RIGHT, fill=Y)
        self.text_callbck_msg.config(yscrollcommand=self.scroll.set)

        self.TabStrip1 = Notebook(self.frame)

        self.tab_control = Notebook(self.TabStrip1)
        # 添加第一个标签页
        self.tab1 = Frame(self.TabStrip1)
        self.tab2 = Frame(self.TabStrip1)

        self.tab_control.add(self.tab1, text='上号结算')

        self.tab2 = Frame(self.TabStrip1)
        self.tab_control.add(self.tab2, text='数据结算')

        self.tab3 = Frame(self.TabStrip1)
        self.tab_control.add(self.tab3, text='广告码处理')

        self.tab4 = Frame(self.TabStrip1)
        self.tab_control.add(self.tab4, text='获取群码')

        self.text_date = Label(self.tab1, text="日期:", background="white")
        # 创建日期选择器
        self.cal = DateEntry(self.tab1, width=12, background='darkblue', foreground='white', borderwidth=2)

        self.start_but = Button(self.tab1, text="选择文件", width=15, command=lambda: self.start_select(1),
                                bg="#3CB371")

        self.loading_but = Button(self.tab1, text="处理文件", width=15, command=lambda: thread_is(self.start_loadning),
                                  bg="#3CB371")

        self.label_port = Label(self.tab2, text="OCR服务地址:", background="white")
        self.text_entry = Entry(self.tab2)

        self.start_but2 = Button(self.tab2, text="选择出粉文件夹", width=15, command=lambda: self.start_select(2),
                                 bg="#3CB371")

        self.start_but3 = Button(self.tab2, text="选择广告码文件夹", width=15, command=lambda: self.start_select(3),
                                 bg="#3CB371")
        self.start_link_but1 = Button(self.tab2, text="选择链接文件", width=15, command=lambda: self.start_select(4),
                                      bg="#3CB371")

        self.loading_but2 = Button(self.tab2, text="处理文件", width=15, command=lambda: thread_is(self.start_work),
                                   bg="#3CB371")

        self.start_link_but = Button(self.tab3, text="选择链接文件", width=15, command=lambda: self.start_select(4),
                                     bg="#3CB371")
        self.text_date1 = Label(self.tab3, text="分裂数量:", width=15, background="white", pady=5)

        self.start_config_but = Button(self.tab3, text="复制图片", width=15, command=lambda: thread_is(self.copy_picture),
                                       bg="#3CB371")
        self.text_entry1 = Entry(self.tab3, width=15, font=('Arial', 17))

        self.generate_but = Button(self.tab3, text="生成广告码", width=15, command=lambda: thread_is(self.start_img_wook),
                                   bg="#3CB371")

        self.generate1_but = Button(self.tab3, text="分裂多张广告码", width=15,
                                    command=lambda: thread_is(self.start_img_generate),
                                    bg="#3CB371")

        self.start_img_but = Button(self.tab3, text="选择图片文件夹", width=15, command=lambda: self.start_select(6),
                                    bg="#3CB371")
        self.start_img_but1 = Button(self.tab3, text="批量修改昵称", width=15,
                                     command=lambda: thread_is(lambda: self.start_img(1)),
                                     bg="#3CB371")
        self.start_img_but2 = Button(self.tab3, text="批量改成女生二维码", width=15,
                                     command=lambda: thread_is(lambda: self.start_img(2)),
                                     bg="#3CB371")

        self.get_img_but = Button(self.tab4, text="一键获取群码", width=15, command=lambda: thread_is(self.start_get_img),
                                  bg="#3CB371")

        self.notebook_size()

        self.text_callbck_msg.bind("<Enter>", lambda event: self.verify_see(event, True, True))
        self.text_callbck_msg.bind("<Leave>", lambda event: self.verify_see(event, False, True))

        self.scroll.bind("<Enter>", lambda event: self.verify_see(event, True, True))
        self.scroll.bind("<Leave>", lambda event: self.verify_see(event, False, True))

    def workbook_init(self, file_name):
        sheet_list = ['日期', '昵称', '状态', '号商', '渠道']
        settle_list = ['渠道', '单价', '补']
        sheet_name = None
        settle_name = None
        df = None
        price_df = None
        # 读取Excel文件
        workbook = load_workbook(file_name, read_only=True)
        # 遍历所有工作表
        for work_name in workbook.sheetnames:
            sheet = workbook[work_name]
            first_row_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            common_elements_count = (Counter(sheet_list) & Counter(first_row_values))
            sheet_count = sum(common_elements_count.values())
            if sheet_count > 2 and sheet_name == None:
                sheet_name = work_name
            common_elements_count = (Counter(settle_list) & Counter(first_row_values))
            settle_count = sum(common_elements_count.values())
            if settle_count > 2 and settle_name == None:
                settle_name = work_name
        if sheet_name:
            df = self.work_book(file_name, sheet_name)
        if settle_name:
            price_df = self.work_book(file_name, settle_name)
        workbook.close()
        gc.collect()  # 触发垃圾回收
        return df, price_df

    def work_result(self, file_name):
        df, price_df = self.workbook_init(file_name)
        # 确保'Date'列是日期格式
        df['日期'] = pd.to_datetime(df['日期'])
        data_time = self.cal.get_date()
        # 查询特定日期的数据，例如查询2023年1月1日的数据
        start_date = data_time
        end_date = data_time
        filtered_df_range = df[(df['日期'] >= pd.to_datetime(start_date)) & (df['日期'] <= pd.to_datetime(end_date))]
        # 按某列分组，例如'ColumnA'
        grouped = filtered_df_range.groupby('号商')
        channel_dict = price_df.set_index('渠道')['单价'].to_dict()
        repair_dict = price_df.set_index('渠道')['补'].to_dict()
        float_dict = price_df.set_index('渠道')['浮动'].to_dict()
        pay_dict = price_df.set_index('渠道')['支付码'].to_dict()
        grouped_dict = price_df.set_index('渠道')['号商'].to_dict()

        df_date = filtered_df_range['日期']
        df_name = filtered_df_range['昵称']
        df_team = filtered_df_range['地推团队']
        df_machine = filtered_df_range['机器']
        df_invite = filtered_df_range['邀请数']
        df_grouped = filtered_df_range['号商']
        df_status = filtered_df_range['状态']
        df_remarks = filtered_df_range['结算备注']
        data_dict = {}

        for date, name, team, machine, invite, grouped, status, remarks in zip(df_date, df_name, df_team, df_machine,
                                                                               df_invite,
                                                                               df_grouped, df_status, df_remarks):
            not_list = ['备', '备用', 'nan', '补', '', 'none']
            team_name = str(team).lower()
            bot_name = machine
            invite_num = invite
            date = date.strftime('%Y-%m-%d')
            dict_key = grouped_dict[grouped]
            if dict_key == None:
                dict_key = grouped
            if remarks == None:
                remarks = ''
            if status == None:
                status = ''
            if name == None:
                name = ''
            if data_dict.get(dict_key):
                data_dict[dict_key].append([date, name, team, machine, invite_num, dict_key, status, remarks])
            else:
                data_dict.update({dict_key: []})
                data_dict[dict_key].append([date, name, team, machine, invite_num, dict_key, status, remarks])
            try:
                bot_name = bot_name.split("机")[0]
            except:
                pass
            if team_name in not_list:
                continue
            if str(invite_num) == 'nan' or invite_num == None:
                invite_num = 0
            if status == '退':
                self.status_dict.update({bot_name: {team_name: int(invite_num)}})
                continue
            if bot_name not in self.boot_dict.keys():
                self.boot_dict.update({bot_name: {team_name: int(invite_num)}})
            else:
                if self.boot_dict[bot_name].get(team_name) != None:
                    if invite_num > self.boot_dict[bot_name].get(team_name):
                        self.boot_dict[bot_name][team_name] = int(invite_num)
                else:
                    self.boot_dict[bot_name].update({team_name: int(invite_num)})
        # 创建Excel文件写入器
        if len(grouped) == 0:
            self.insert_text("没有找到当天数据", 'red')
            return
        # 创建一个Workbook对象，这将会创建一个新的Excel文件
        wb = Workbook()
        wb_verify = True
        # 获取当前活跃的worksheet
        for data_key in data_dict:
            data_itme = data_dict[data_key]
            if wb_verify:
                # 获取当前活跃的worksheet
                ws = wb.active
                ws.title = data_key
                wb_verify = False
            else:
                # 更改worksheet的名称
                ws = wb.create_sheet(data_key)
            complete_count = 0
            repair_count = 0
            otehr_count = 0
            half_count = 0
            price = channel_dict[data_key]
            pay = pay_dict[data_key]
            repair = repair_dict[data_key]
            float_prce = json.loads(float_dict[data_key])
            category = str(data_key)
            sorted_reversed_list = [int(x) for x in reversed(sorted(list(float_prce.keys())))]
            sorted_reversed_list.sort(reverse=True)
            if float_prce != {}:
                for item in sorted_reversed_list:
                    if complete_count + repair_count >= item:
                        price = price + float_prce[str(item)]
                        break
            # 添加表头
            ws.append(['日期', '昵称', '号商', '状态', '结算备注'])
            # 例如，设置列宽和行高（可选）
            for column in "ABCDE":  # 根据需要调整列宽和行高，例如A、B、C列
                if 'E' in column:
                    ws.column_dimensions[column].width = 30  # 设置列宽为20个字符宽
                else:
                    ws.column_dimensions[column].width = 20  # 设置列宽为20个字符宽
                font = Font(color=Color(rgb='000000'))  # 红色字体
                # 设置填充颜色
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell = ws[column][0]
                cell.font = font
                cell.fill = fill
            for elem in data_itme:
                if '结一半' in elem[-1]:
                    half_count += 1
                if elem[-2] == '打完' or elem[-2] == '用完':
                    complete_count += 1
                    self.complete_number += 1
                elif elem[-2] == '补':
                    repair_count += 1
                    self.repair_number += 1
                elif elem[-2] == '退':
                    continue
                else:
                    otehr_count += 1
                    self.otehr_number += 1
                ws.append(elem[:2] + elem[5:])
            # 删除为空的子表
            if complete_count == 0 and repair_count == 0 and otehr_count == 0:
                wb.remove(wb[data_key])
                continue
            # 打完
            price_count = complete_count * price
            price_text = f'{complete_count}x{price}={price_count}'
            if repair_count != 0:
                price_count = repair_count * repair + price_count
                price_text = f'{complete_count}x{price}+{repair_count}x{repair}={price_count}'
            row_data = ['', '', f'共计：{complete_count + repair_count + otehr_count}',
                        f'用{complete_count}补{repair_count}', '', '']
            channel_text = f'共计:{complete_count + repair_count + otehr_count},用:{complete_count},补:{repair_count},结算:{price_text}'
            self.price_count += price_count
            ws.append(row_data)
            ws.append(['', '', '', '', f'{price_text}', ''])
            if half_count != 0:
                half_price = int(half_count * price / 2)
                ws.append(
                    ['', '', '', '', f'{price_count}-{half_count}x{int(price / 2)}={price_count - half_price}', ''])
                self.price_count -= half_price
                self.channel_list.append(
                    [f"{category},{channel_text}-{half_count}x{int(price / 2)}={price_count - half_price}",
                     price_text])
            else:
                self.channel_list.append([f"{category},{channel_text}", price_text])
            try:
                if pay != None:
                    pay_file = os.path.join(os.path.dirname(self.file_path), '支付码')
                    # 添加图片到工作表，'path_to_image'是你的图片文件路径，'H1'是图片放置的起始单元格
                    img = OpenpyxlImage(os.path.join(pay_file, pay))
                    img.width = 400
                    img.height = 600
                    ws.add_image(img, 'H1')
            except:
                pass
            # 设置所有单元格文字居中
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        # 创建一个Workbook对象，这将会创建一个新的Excel文件
        result_file_name = os.path.join(os.path.dirname(file_name), f'{data_time}结算记录.xlsx')
        # 保存文件
        wb.save(result_file_name)
        os.startfile(os.path.dirname(file_name))

    def crop_img_num(self, img, box):
        x1, y1, x2, y2 = box
        crop = img.crop((x1[0] - 5, y1[1] - 100, x2[0] + 5, y2[1] + 10))
        crop_rgb = crop.convert('RGB')
        buffered = io.BytesIO()
        crop_rgb.save(buffered, format="JPEG")
        image_bytes = buffered.getvalue()
        # 将字节流转换为Base64编码的字符串
        base64_str = base64.b64encode(image_bytes).decode('utf-8')
        result_text = self.ocr_data(base64_str)
        data_text = result_text.get("data").replace('\n', '').replace(' ', '')
        num = 0
        try:
            match = re.search(r'\d+', data_text)
            num = match.group()
        except:
            crop = img.crop((x1[0] - 5, y1[1] - 35, x2[0] + 5, y2[1] - 20))
            crop_rgb = crop.convert('RGB')
            # 转换为灰度图并增强对比度
            crop_rgb = crop_rgb.convert('L')
            enhancer = ImageEnhance.Contrast(crop_rgb)
            crop_rgb = enhancer.enhance(2.0)  # 增强对比度
            buffered = io.BytesIO()
            crop_rgb.save(buffered, format="JPEG")
            image_bytes = buffered.getvalue()
            # 将字节流转换为Base64编码的字符串
            base64_str = base64.b64encode(image_bytes).decode('utf-8')
            result_text = self.ocr_data(base64_str)
            data_text = result_text.get("data").replace('\n', '').replace(' ', '')
            try:
                match = re.search(r'\d+', data_text)
                num = match.group()
            except Exception as e:
                self.insert_text(f'{e}', 'red')
        return num

    def crop_img_hu(self, img):
        img_rgb = img.convert('RGB')
        buffered = io.BytesIO()
        img_rgb.save(buffered, format="JPEG")
        image_bytes = buffered.getvalue()
        # 将字节流转换为Base64编码的字符串
        base64_str = base64.b64encode(image_bytes).decode('utf-8')
        result = self.ocr_data(base64_str, "dict")
        result_data = result.get('data')
        cawcdeb = None
        box = None
        link_box = None
        name = None
        verify = True
        for item in result_data:
            text = item['text']
            if 'http' in text:
                verify = False
            if verify:
                match = re.search(r"\dGl", text)
                matchs = re.search(r"^\d$", text)
                if match == None and matchs == None:
                    name = text
            if 'cawcde' in text:
                cawcdeb = f'cawcde{text.split("cawcde")[1]}'
            elif 'awcde' in text:
                cawcdeb = f'cawcde{text.split("awcde")[1]}'
            d = difflib.SequenceMatcher(None, text, "新增客户")
            ratio = d.ratio() * 100
            if '新增客户' in text:
                box = item['box']
            elif ratio > 60:
                box = item['box']
            d = difflib.SequenceMatcher(None, text, "打开链接")
            ratio = d.ratio() * 100
            if '打开链接' in text:
                link_box = item['box']
            elif ratio > 60:
                link_box = item['box']

        num = self.crop_img_num(img, box)
        link_num = self.crop_img_num(img, link_box)
        return name, cawcdeb, num, link_num

    def img_file_time(self, path):
        while True:
            if os.path.isdir(path):
                path_dirname = os.listdir(path)[0]
                path = os.path.join(path, path_dirname)
            else:
                break
        return path

    def img_data(self, path_dirname):
        # img_hu=os.path.join(path_dirname, '湖')
        img_list = os.listdir(path_dirname)
        dict_hu = {}
        self.cawcdeb_list = []
        for item in img_list:
            img_file = os.path.join(path_dirname, item)
            if is_image(img_file):
                self.cawcde_num += 1
                # 打开图像文件
                image = Image.open(img_file)
                # 例如，我们想截取前100像素的高度，宽度与原图相同
                # height_to_crop = image.height / 2
                # width_to_crop = image.width  # 或者你想要的任何宽度
                # cropped_image = image.crop((0, 0, width_to_crop, height_to_crop))
                result = self.crop_img_hu(image)
                name = result[0]
                cawcdeb = result[1]
                num = result[2]
                link_num = result[3]
                self.cawcdeb_list.append(cawcdeb)
                self.insert_text(f'{name}\t{cawcdeb}\t访问量:{link_num}\t{num}个粉', 'green', f'{num}个粉')
                dict_hu.update({cawcdeb: {"name": name, "num": int(num), "link_num": int(link_num)}})
                new_name = os.path.join(path_dirname, f"{cawcdeb}.jpg")
                try:
                    os.rename(img_file, new_name)
                except Exception as e:
                    self.insert_text(f'{e}', 'red')
        return dict_hu

    def is_valid_url(self, url):
        try:
            headers = {
                "Content-Type": "text/html; charset=utf-8",
                "User-Agent": 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1',

            }
            result = urlparse(url)
            if all([result.scheme, result.netloc]):  # 基本URL格式验证
                response = requests.head(url, allow_redirects=True, timeout=5, headers=headers)  # 发送HEAD请求以检查URL是否可达
                if response.status_code == 200:  # 返回200表示URL有效且可达
                    return False
            else:
                return True
        except requests.RequestException as e:  # 捕获请求异常或解析异常
            self.insert_text(f"{e}", 'red')
            return True

    def start_img_generate(self):
        try:
            if self.file_text or self.folder_img:
                if self.file_text:
                    cawcdeb_list = open(self.file_text, 'r').readlines()
                    code_list = [item.replace("\n", "") for item in cawcdeb_list]
                    counter = Counter(code_list)
                    duplicates = [item for item, count in counter.items() if count > 1]
                    if duplicates and duplicates != ['']:
                        self.insert_text(f"重复码", 'red')
                        for item in duplicates:
                            self.insert_text(f"{item}")
                        self.insert_text(f"共计：{len(duplicates)}")
                    dict_list = list(set(code_list))
                    dict_list.sort()
                    base_dirname = os.path.dirname(self.file_text)
                    base_name = os.path.basename(self.file_text)
                    dir_name = base_name.split('.')[0]
                else:
                    dict_list = os.listdir(self.folder_img)
                    base_dirname = os.path.dirname(self.folder_img)
                    dir_name = os.path.basename(self.folder_img)
                fission_count = self.text_entry1.get()
                name_coun = 1
                count = 0
                for dict_itme in dict_list:
                    dir_path = os.path.join(base_dirname, dir_name + str(name_coun))
                    os.makedirs(dir_path, exist_ok=True)
                    img_count = 0
                    for i in range(0, int(fission_count)):
                        if self.file_text:
                            # 提取字符中的链接
                            pattern = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
                            urls = re.findall(pattern, dict_itme)
                            if urls:
                                url = urls[0]
                                # 判断链接是否有效
                                if self.is_valid_url(url):
                                    self.insert_text(f"请检查链接是否有误\t{url}", 'red')
                                    return
                                cawcdeb = url.split("/")[-1]
                                img_file = os.path.join(dir_path, f'{cawcdeb}.png')
                                if bool(re.match(r'^[^\\\/:*?"<>|\r\n]+$', f'{cawcdeb}.png')) == False:
                                    cawcdeb = url.split("?s=")[0].split("/")[-1]
                                    if bool(re.match(r'^[^\\\/:*?"<>|\r\n]+$', f'{cawcdeb}.png')):
                                        img_file = os.path.join(dir_path, f'{cawcdeb}.png')
                                    else:
                                        img_file = os.path.join(dir_path, f'{name_coun}.png')
                                img_name = img_file.replace(cawcdeb, f'{i + 1}-{cawcdeb}')
                                result = self.qrcode_img(url, img_name)
                                if result:
                                    return
                        else:
                            img_file = os.path.join(self.folder_img, dict_itme)
                            image = Image.open(img_file)
                            decoded_objects = decode(image)
                            if decoded_objects:
                                url = decoded_objects[0].data.decode('utf-8')
                                img_name = os.path.join(dir_path, f'{i + 1}-{dict_itme}')
                                result = self.qrcode_img(url, img_name)
                                if result:
                                    return
                            else:
                                self.insert_text(f"{dict_itme}此图片识别识别", 'red')
                                return

                        img_count += 1
                        count += 1
                    self.insert_text(f"{dir_name + str(name_coun)}共:{img_count}张", 'green')
                    name_coun += 1
                self.conf.set_option("Setting", "fission", fission_count)
                self.conf.save_config()
                self.insert_text(f"共:{count}张", 'green')
            else:
                self.insert_text(f"请选择分裂txt文件或者图片文件夹", 'red')
        except Exception as e:
            self.insert_text(f"{e}", 'red')

    def start_img_wook(self):
        try:
            if self.file_text:
                cawcdeb_list = open(self.file_text, 'r').readlines()
                code_list = [item.replace("\n", "") for item in cawcdeb_list]
                counter = Counter(code_list)
                duplicates = [item for item, count in counter.items() if count > 1]
                if duplicates and duplicates != ['']:
                    self.insert_text(f"重复码", 'red')
                    for item in duplicates:
                        self.insert_text(f"{item}")
                    self.insert_text(f"共计：{len(duplicates)}")
                dict_list = list(set(code_list))
                dict_list.sort()
                base_dirname = os.path.dirname(self.file_text)
                base_name = os.path.basename(self.file_text)
                dir_name = base_name.split('.')[0]
                dir_path = os.path.join(base_dirname, dir_name)
                os.makedirs(dir_path, exist_ok=True)
                self.insert_text(f"{dir_path}", 'green')
                count = 0
                for item in dict_list:
                    # 提取字符中的链接
                    pattern = r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+'
                    urls = re.findall(pattern, item)
                    if urls:
                        url = urls[0]
                        # 判断链接是否有效
                        if self.is_valid_url(url):
                            self.insert_text(f"请检查链接是否有误\t{url}", 'red')
                            return
                        img_file = os.path.join(dir_path, f'{url.split("/")[-1]}.png')
                        if bool(re.match(r'^[^\\\/:*?"<>|\r\n]+$', f'{url.split("/")[-1]}.png')) == False:
                            if bool(re.match(r'^[^\\\/:*?"<>|\r\n]+$', f'{url.split("?s=")[0].split("/")[-1]}.png')):
                                img_file = os.path.join(dir_path, f'{url.split("?s=")[0].split("/")[-1]}.png')
                            else:
                                img_file = os.path.join(dir_path, f'{count}.png')
                        result = self.qrcode_img(url, img_file)
                        if result:
                            return
                        count += 1
                self.insert_text(f"共:{count}张", 'green')
                os.startfile(dir_path)
            else:
                self.insert_text(f'请选择txt链接文件', 'red')
        except Exception as e:
            self.insert_text(f'{e}', 'red')

    def qrcode_img(self, url, img_file):

        # 获取当前执行的 exe 文件的路径
        # exe_path = sys.executable
        # 获取 exe 文件所在的目录
        # exe_dir = os.path.dirname(exe_path)
        # self.insert_text(f"选择默认地址 {self.folder_config}", 'red')
        self.folder_config = os.path.join(os.getcwd(), 'config')
        if os.path.exists(self.folder_config):
            # 创建一个白色背景的图像
            width, height = 686, 940
            image = Image.new('RGB', (width, height), 'white')
            draw = ImageDraw.Draw(image)
            icon_flie = os.path.join(self.folder_config, 'icon')
            icon_list = os.listdir(icon_flie)
            icon_name = os.path.join(icon_flie, icon_list[random.randint(0, len(icon_list) - 1)])
            icon_x = 100
            icon = self.create_rounded_icon(icon_name)
            # 粘贴图标到背景
            image.paste(icon, (icon_x, 100), icon)
            # 加载并调整女性图标
            female_img = os.path.join(self.folder_config, 'female.png')
            female = Image.open(female_img).convert("RGBA")  # 确保图标是RGBA模式
            female = female.resize((25, 25))  # 调整图标大小
            # 添加文字
            ttf_flie = os.path.join(self.folder_config, 'SimHei.ttf')
            name_font = ImageFont.truetype(ttf_flie, 30)  # 确保有arial.ttf字体文件
            area_font = ImageFont.truetype(ttf_flie, 25)  # 确保有arial.ttf字体文件
            name_text = os.path.join(self.folder_config, 'name.txt')
            name_list = open(name_text, encoding='utf-8').readlines()
            name = name_list[random.randint(0, len(name_list) - 1)]
            name_width, name_height = draw.textsize(name, font=name_font)
            name_x = 220  # 从左边固定位置开始
            draw.text((name_x, 110), name, font=name_font, fill='black')
            # 根据name的位置调整female的位置
            female_x = name_x + name_width + 2  # 在name后面5像素的位置
            image.paste(female, (int(female_x), 113), female)
            area = '中国'
            draw.text((name_x, 160), area, font=area_font, fill='grey')
            text = "扫一扫上面的二维码图案，加我为朋友。"
            text_width, text_height = draw.textsize(text, font=area_font)
            text_x = (width - text_width) / 2
            text_y = height - text_height - 130
            draw.text((text_x, text_y), text, font=area_font, fill='lightgray')

            qrcode_text = '该二维码已通过官方检测，安全可用'
            draw.textsize(qrcode_text, font=area_font)
            draw.text((text_x + 10, text_y + 50), qrcode_text, font=area_font, fill='lightgray')

            # overlay = self.qrcode_make(url)
            position = (icon_x - 3, 250)
            target_size = (480, 480)
            overlay = qrcode.make(data=url)
            overlay = overlay.resize(target_size)
            # 将贴图贴到背景图片上
            image.paste(overlay, position)
            # image.show()
            # 保存结果图片
            # 保存图像
            image.save(img_file)
            return False
        else:
            self.insert_text(f"请选择配置文件", 'red')
            return True

    def copy_files_to_clipboard(self, file_paths):
        """
        将文件路径列表复制到剪贴板，实现类似“复制文件”操作。

        参数:
          file_paths: 文件路径字符串列表（需要使用绝对路径）
        """
        # 构造文件列表字符串：
        # 每个文件路径后面以 '\0' 结束，最后再添加一个额外的 '\0'
        file_list = "\0".join(file_paths) + "\0\0"
        # Windows 剪贴板要求使用 Unicode 格式（utf-16le 编码）
        file_list_bytes = file_list.encode("utf-16le")

        # 构造 DROPFILES 结构体
        # 结构体定义:
        # typedef struct _DROPFILES {
        #   DWORD pFiles;   // 从结构体起始到文件列表数据的偏移字节数，一般为20
        #   POINT pt;       // 拖放时的坐标（这里设为0）
        #   BOOL fNC;       // 非客户区标志（设为0）
        #   BOOL fWide;     // 是否为 Unicode 格式，非0表示 Unicode（设为1）
        # } DROPFILES;
        #
        # 使用 struct.pack 进行打包，注意使用小端格式 "<"
        dropfiles = struct.pack("<IiiII", 20, 0, 0, 0, 1)
        # 将结构体和文件列表数据拼接
        data = dropfiles + file_list_bytes
        # 打开并清空剪贴板，然后设置 CF_HDROP 数据格式（用于文件拖放操作）
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32con.CF_HDROP, data)
        # 同时设置“Preferred DropEffect”格式，指定复制（1：复制；0：剪切）
        cf_drop_effect = win32clipboard.RegisterClipboardFormat("Preferred DropEffect")
        win32clipboard.SetClipboardData(cf_drop_effect, struct.pack("<I", 1))
        win32clipboard.CloseClipboard()

    def on_ctrl_v(self):
        try:
            img_list = self.boot_list[self.slice_count: self.slice_count + self.max]
            if img_list:
                self.insert_text([x.split('\\')[-1].split('-')[0] for x in img_list])
                self.copy_files_to_clipboard(img_list)
                self.slice_count += self.max
            else:
                self.insert_text("没有图片了", 'red')
                # 清空剪切板
                pyperclip.copy('')
                keyboard.remove_hotkey("ctrl+v")
                return
        except Exception as e:
            self.insert_text(f"{e}", 'red')

    def copy_picture(self):
        def get_all_images(parent_dir, extensions=('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff')):
            """
            获取指定父目录下所有图片文件
            :param parent_dir: 目标父目录路径
            :param extensions: 支持的图片扩展名元组
            :return: 排序后的图片路径列表
            """
            parent_path = Path(parent_dir).resolve()
            if not parent_path.exists():
                raise FileNotFoundError(f"目录不存在: {parent_path}")
            image_files = []
            for root, _, files in os.walk(parent_path):
                for file in files:
                    if file.lower().endswith(extensions):
                        file_address = os.path.join(root, file)
                        image_files.append(file_address)
            return sorted(image_files)

        try:
            self.slice_count = 0
            self.max = 9
            try:
                keyboard.remove_hotkey("ctrl+v")
            except:
                pass
            self.insert_text('按下ctrl+v键开始复制', 'green')


        except Exception as e:
            self.insert_text(f'{e}', 'red')

        def sort_image_files(image_paths):
            # 自定义排序规则：按文件夹编号和数字部分升序排序
            def extract_key(filename):
                # 提取路径中的文件夹编号（如S10中的10）
                folder_part = filename.split('\\')[-1].split('-')[0]
                folder_num = int(folder_part[1:]) if folder_part.startswith('S') else 0
                return folder_num

            # 使用sorted函数进行排序
            sorted_paths = sorted(image_paths, key=extract_key)
            return sorted_paths

        if self.folder_img:
            images = get_all_images(self.folder_img)
            # 将数字 字符进行排序
            self.boot_list = sort_image_files(images)
            img_list = self.boot_list[self.slice_count: self.slice_count + self.max]
            self.insert_text([x.split('\\')[-1].split('-')[0] for x in img_list])
            self.copy_files_to_clipboard(img_list)
            self.slice_count += self.max
            try:
                keyboard.add_hotkey('ctrl+v', self.on_ctrl_v)
            except Exception as e:
                self.insert_text(f"操作失败: {str(e)}", 'red')
        else:
            self.insert_text("请选择需要复制的图片文件夹", 'red')


root = Tk()


def close():
    root.destroy()


root.protocol("WM_DELETE_WINDOW", close)
app = Application(master=root)

if __name__ == '__main__':
    try:
        app.mainloop()
    except Exception as e:
        print("发生异常:", e)
        root.destroy()  # 退出主循环，关闭所有进程
